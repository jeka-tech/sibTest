

import os, config, pandas as pd
from openpyxl import load_workbook, writer
from numpy import*
data = []

os.chdir(config.project_path)       # переходим в корневую директорию проекта

# активируем ексель-файл рабочую книгу с рабочим Листом вопросов-ответов Теста
workbook = load_workbook('.\\xls_works\\testPRG_vopr_db.xlsx')
worksheet = workbook['db_vopr']


# непонятная мне ботва, незнаю откуда взялась, надо выяснить, удалять не стал пока, хотя без нее вроде рвботает
cwd = os.getcwd()
cwd


# from numpy import *
#0 x = [2]
#0 y = ["vfyufyufyuvuyv"]
#0 z = [4]


# A = arange(y*x).reshape(x,y)
# A = ones(x*y).reshape(x,y)

#0 D = bmat([[[x]], [[y]], [[z]]])
#0 print(D)


#0 A = bmat([[y]])
#0 print(A)

#0 C = bmat([[z]])

# A[0] = 1
#0 B = []
#0 B.append(A)
#0 print(B)
#0 B.append(C)
#0 print(B)

#0 E = bmat([[[]]])
#0 print(E)
#0 C = []
#0 C.append(E)
#0 print(C)
c = 0
# cell_value = 0
# Блок преброски- переноса данных из ексель-файла из активированного выше Листа в массив данных и запись их во фрейм
res_ans = ("№ ячеек верных ответов")   # заведем переменную для списка верных ответов в строке и сразу запишем загловок
# res_ans.append("№ ячеек верных ответов")
for row in worksheet.rows:                              # перебираем строки в файле
    row_cells = []                                      # заведем переменную для списка ячеек в строке
                                                      # заведем переменную для списка верных ответов в строке
    c+=1
    #0 G = bmat([[[]]])
    if row[1].value:                                    # если ячейка с вопросом не пустая (второй столбец имеет индекс 1), то...
        for cell in row:                                # перебираем все ячейки в этой строке
            # print(row_cells)
            # if not(row_cells == 'None'):
            #     continue
            row_cells.append(cell.value)                # и добавляем в список значения ячеек в этой строке
            #0 G = [cell.value]
            #0 C.append (G)
        for cell in row:                                # повторно перебираем все ячейки в этой строке
            if cell.fill.fgColor.rgb != '00000000':     # чтобы проверить, какая ячейка (ячейки) имеют заливку, отличную от белой (заливкой выделен правильный ответ)
                res_ans.append(cell.column - 2)          # и добавляем во вложенный список для фиксации ячеек с правильным ответом

        row_cells.insert(2, res_ans)    # добавляем заголовок в верхнюю ячейку правильного(ых) варианта(ов) ответа(ов) в начало списка вопросов

    # # код избавления от фразы list с преобразованием в кортеж но только запятая после переменной появляется
       #  if c == 1 :
       #      row_cells.insert(2, res_ans)    # добавляем заголовок в верхнюю ячейку правильного(ых) варианта(ов) ответа(ов) в начало списка вопросов
       #  else:
       #      cell_value = tuple(res_ans)     #  преобразование списка в кортеж
       #      row_cells.insert(2, cell_value)  # добавляем кортеж правильного(ых) варианта(ов) в ячейки последовательно сверху вниз

        # row_cells.insert(2, res_ans)    # добавляем список правильного(ых) варианта(ов) ответа(ов) в начало списка вопросов

        # # код избавления от фразы list через перевод в строковое
        # if c == 1 :
        #     row_cells.insert(2, res_ans)    # добавляем список правильного(ых) варианта(ов) ответа(ов) в начало списка вопросов
        # else:
        #     cell_value = str(res_ans[0])
        #     for i in range(1, len(res_ans)):
        #         cell_value += ', ' + str(res_ans[i])
        #     row_cells.insert(2, cell_value)




    # print(row_cells)

    # Код избавления от None
    if row_cells == []:
        continue
    # print(row_cells)
    data.append(row_cells)                              # и добавляем полученный список в основной массив данных
    res_ans = []                                        # обнуляем переменную списка верных ответов в строке

df = pd.DataFrame(data)     # заведем и запишем в переменную в виде фрейма таблицу данных полученных из ексель-файла

#    E = [row_cells]
#    C.append(E)


#0 print(""" ""
    #0   "ыыы"
#0  "" """ )

#0 print(G)

#0 print(C)



# workbook = load_workbook('.\\xls_works\\testPRG_vopr_db.xlsx')
# writer = pd.ExcelWriter('.\\xls_works\\testPRG_vopr_db.xlsx')
# writer.workbook = workbook
# df.to_excel(writer, 'Sheet2')
# writer.save()



# df2 = []
#1
#1 workbk = load_workbook("example.xlsx")

# для промежуточной записи в обезличенной форме массива данных активируем ексель-файл с новым Лист2 вопросов-ответов Теста
# Указать writer библиотеки
writer = pd.ExcelWriter('.\\xls_works\\testPRG_vopr_db.xlsx')
writer._book = workbook
# df.to_excel(df2, sheet_name= 'Sheet2', index=False)
# Сохраним результат
# df.save()
df.to_excel(writer, sheet_name='Sheet2', index=False)
# index=False
#     ExcelWriter = pd.ExcelWriter("example.xlsx")
writer._save()


# writer._save(showHeaders=False)

#2 df.to_excel("example.xlsx", sheet_name= 'Sheet1', index=False)
# worksh = workbook['Sheet2']
# print('печать 1')
# print(worksh)



#3 подготовка к сравнению и сранение ранее сформированной базы данных в ексель файле итоговой с изменениями в корневой
df_kornres = df
data_itog = []   # заведем переменную для матрицы итоговых коненых данных

exfile_itog = load_workbook('.\\xls_works\\testPRG_vopr_db.xlsx')
itog_sheet = exfile_itog ['Sheet2']
worksheet = exfile_itog['Sheet2']
# print(worksheet)
worksheet.delete_rows(1, amount=1)
exfile_itog.save('.\\xls_works\\testPRG_vopr_db.xlsx')
# itog_sheet = exfile_itog ['Sheet2']

# создаем из листа 2 новый фрейм итоговых данных после промежуточных сохраненных изменений из первичного фрейма
# df2 = pd.read_excel(      '.\\xls_works\\testPRG_vopr_db.xlsx', sheet_name='Sheet2')
# print('печать 2')
# print(df2)
# for index, row in itog_sheet.iloc[1:].iterrows():    # перебираем строки в файле
for row in itog_sheet.rows:                              # перебираем строки в файле
    row_cells = []                                      # заведем переменную для списка ячеек в строке
    # print('печать row')
    # print(row[0])
#    res_ans = []                                        # заведем переменную для списка верных ответов в строке
    if row[1].value:                                     # если ячейка с вопросом не пустая (второй столбец имеет индекс 1), то...
       for cell in row:                                # перебираем все ячейки в этой строке
           row_cells.append(cell.value)                # и добавляем в список значения ячеек в этой строке
            # print(cell)
    data_itog.append(row_cells)                              # и добавляем полученный список в основной массив данных
    #     res_ans = []                                        # обнуляем переменную списка верных ответов в строке
df_itog = pd.DataFrame(data_itog)



#3 print(data_itog)
#3 print(itog_sheet)
# print(df_kornres.info)
# print(df_itog.info)

# создание для проверки фрейма сравнения в целях эксперимента с командой прямого сравнения
#СРАВН  df_comp = df_kornres.compare(df_itog)
# создание матрицы данных из фрейма сраванения
#СРАВН  ttm = df_comp.values

# создание матриц данных из фреймов первичного-корневого листа данных 2 и итогового промежуточного листа 3 к будущим сравнениям
#СРАВН  ttm_korn = df_kornres.values
#СРАВН  ttm_itog = df_itog.values
# ttm_korn = df_itog.values
ttm_korn = df.values

# запишем в отдельный лист 3 в обезличенной форме фрейм первичного листа данных
#СРАВН  workbook = load_workbook('.\\xls_works\\testPRG_vopr_db.xlsx')
#СРАВН  writer = pd.ExcelWriter('.\\xls_works\\testPRG_vopr_db.xlsx')
#СРАВН  writer._book = workbook
#СРАВН  df.to_excel(writer, sheet_name='Sheet3', index=False)
#СРАВН  writer._save()

# так как при записи в обезличенной форме фрейма первичного листа данных появляется лишняя верхняя строка с номерами
# то надо ее удалить из этого листа 3, активируем лист и удаляем лишнюю строку, активный лист используется далее
#СРАВН  exfile_korn = load_workbook('.\\xls_works\\testPRG_vopr_db.xlsx')
#СРАВН  korn_sheet = exfile_korn ['Sheet3']
#СРАВН  worksheet = exfile_korn['Sheet3']
# print(worksheet)
# удаляем лишнюю строку
#СРАВН  worksheet.delete_rows(1, amount=1)
#СРАВН  exfile_korn.save('.\\xls_works\\testPRG_vopr_db.xlsx')


# создаем из листа 3 новый фрейм корневых первичных данных, но уже с листа 3 из обезличенной формы и без лишней строки
#СРАВН  korn_data = []
#СРАВН  for row in korn_sheet.rows:                              # перебираем строки в файле
#СРАВН      row_cells = []                                      # заведем переменную для списка ячеек в строке
    # print('печать row')
    # print(row[0])
#    res_ans = []                                        # заведем переменную для списка верных ответов в строке
#СРАВН      if row[1].value:                                     # если ячейка с вопросом не пустая (второй столбец имеет индекс 1), то...
#СРАВН          for cell in row:                                # перебираем все ячейки в этой строке
#СРАВН              row_cells.append(cell.value)                # и добавляем в список значения ячеек в этой строке
            # print(cell)
#СРАВН      korn_data.append(row_cells)                              # и добавляем полученный список в основной массив данных
    #     res_ans = []                                        # обнуляем переменную списка верных ответов в строке
#
# создаем "причесанный" фрейм первичных данных
#СРАВН  df_korn = pd.DataFrame(korn_data)
# создаем матрицу данных из "причесанного" фрейма первичных итоговых данных
#СРАВН  ttm_korn = df_korn.values

# расчпечатываем полученные матрицы, чтобы убедиться в отсутствии лишних пояснений перед  вложенными в них списками
#СРАВН  print(ttm_korn)

#СРАВН  print('''


#СРАВН         итого
#СРАВН         ''')


#СРАВН  print(ttm_itog)

# делаем сравнение полученных матриц командой прямого сравнения матриц и печатаем результат сравнения
#СРАВН  result = array_equal(ttm_korn, ttm_itog)
#СРАВН  print(result)

# if df_kornres.equals(df_itog):
# if ttm_korn == ttm_itog:
    # print('без изменений')
    # else:
    # print('поплыло')



if __name__ == '__main__':
    # print(data)
    print(ttm_korn)

    print('конец')
    # print(df)
#   pass